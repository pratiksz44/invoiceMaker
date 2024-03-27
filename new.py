import tkinter
import shutil
from tkinter import ttk
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment,Side,Border
import datetime
from num2words import num2words
#import win32com.client
import tkinter.messagebox 


invoice_list = []

def caarItem():
    qty_entry.delete(0, tkinter.END)
    qty_entry.insert(0,'')
    price_entry.delete(0, tkinter.END)
    price_entry.insert(0,'')
    desc_entry.delete(0,tkinter.END)
    desc_entry.insert(0,'')

def add_item():
    qty = int(qty_entry.get())
    price =float(price_entry.get())
    desc = desc_entry.get()  
    total = price * qty
    invoice_item = [desc,qty,price,total]
    tree.insert('',0,values=invoice_item)
    invoice_list.append([desc,qty,price,total])
    caarItem()

def delete():
    try:
        seleI = tree.selection()[0]
        tree.delete(seleI)
        index = int(seleI[2]+seleI[3])
        index = index -1
        invoice_list.pop(index)
        print(invoice_list)
    except(Exception):
         print(Exception)

def newI():
    caarItem()
    cNo_entry.delete(0,tkinter.END)
    cNo_entry.insert(0,'')
    tree.delete(*tree.get_children())
    invoice_list.clear()

def save():
    if True:
        #ws.cell(row=6,column=7).value=None
        src = 'temp.xlsx'
        cn = cNo_entry.get()
        cnn = cn
        cn = cn +".xlsx"
        shutil.copy(src,cn)
        wb = openpyxl.load_workbook(cn)
        ws = wb.get_sheet_by_name('INVOICE')
        ccn = 'INVOICE NO :- '+cn
        ws.cell(row=7,column=7).value= cnn
        now =str( datetime.date.today())
        dte = ' DATE :- '+now
        ws.cell(row=8,column=7).value = dte
        ws.cell(row=9,column=7).value = 'Delivery Challan No. '+cnn+dte
        poN = po_entry.get()
        po = 'P.O. NO.:- '+poN
        pd = poD_entry.get()
        pdate = 'Date:-    '+pd
        ws.cell(row=10,column=7).value = po
        ws.cell(row=11,column=7).value = pdate
        start = 24
        x = 1;
        for d in invoice_list:
            ws.cell(row=start,column=1).value = x;
            x += 1
            ws.cell(row=start,column=2).value = d[0];
            ws.cell(row=start,column=6).value = d[1];
            ws.cell(row=start,column=7).value = d[1];
            ws.cell(row=start,column=8).value = d[2];
            ws.cell(row=start,column=9).value = d[3];
        subtotal = sum(i[3] for i in invoice_list)
        ws.cell(row=41,column=9).value = subtotal;
        gt = int(gst_entry.get())
        p = float(gt/100)
        tax = subtotal*p
        ws.cell(row=42,column=9).value = tax;
        ws.cell(row=43,column=9).value = tax;
        ws.cell(row=44,column=9).value = tax*2;
        ttl = (tax*2)+subtotal
        ws.cell(row=49,column=9).value = ttl;
        tiw = num2words(ttl, to = 'ordinal')
        ws.cell(row=47,column=1).value = tiw;
        wb.save(cn)
    else:
        tkinter.messagebox.showinfo("Error",  "Fill all the fields") 


    
# def printI():
#     save()
#     excel = win32com.client.Dispatch("Excel.Application")
#     excel.Visible = True
#     filename = cNo_entry.get()+'.xlsx'
#       # Optional: Set to True if you want Excel to be visible
#     workbook = excel.Workbooks.Open(filename)
    
#     try:
#         sheet = workbook.Worksheets('INVOICE')
#         sheet.PrintOut()
#     except Exception as e:
#         print(f"Error printing sheet: {e}")
#     finally:
#         workbook.Close(False)
#         excel.Quit()
    



window = tkinter.Tk()
window.title("Raj Engg Invoice Generator")

x = tkinter.Label(window,text="RAJ ENGINEERING",font=50)
x.pack()

frame = tkinter.Frame(window)
frame.pack()



cNo = tkinter.Label(frame ,text="Add challan No :")
cNo.grid(row=1,column=0)
cNo_entry = tkinter.Entry(frame)
cNo_entry.grid(row=1,column=1)

po = tkinter.Label(frame, text="Add PO No. :")
po.grid(row=1,column=2)
po_entry = tkinter.Entry(frame)
po_entry.grid(row=1,column=3)

poD = tkinter.Label(frame, text="Add PO Date :")
poD.grid(row=1,column=4)
poD_entry = tkinter.Entry(frame)
poD_entry.grid(row=1,column=5)


desc = tkinter.Label(frame ,text="Add description :")
desc.grid(row=2,column=0)
desc_entry = tkinter.Entry(frame)
desc_entry.grid(row=2,column=1)

qty = tkinter.Label(frame , text="Enter qty :")
qty.grid(row=2,column=2)
qty_entry = tkinter.Entry(frame)
qty_entry.grid(row=2,column=3)

price = tkinter.Label(frame,text="Enter price :")
price.grid(row=2,column=4)
price_entry = tkinter.Entry(frame)
price_entry.grid(row=2,column=5)

add_btn = tkinter.Button(frame,text="Add item",command=add_item)
add_btn.grid(row=3,column=0)

del_btn = tkinter.Button(frame,text="Delete item", command=delete)
del_btn.grid(row=3,column=1)

gst = tkinter.Label(frame,text="Enter GST :")
gst.grid(row=3,column=2)
gst_entry = tkinter.Entry(frame)
gst_entry.grid(row=3,column=3)


cols = ('Description','Qty','Price','Total')
tree = ttk.Treeview(frame,columns=cols,show="headings")
tree.grid(row=4,column=0,columnspan=6,padx=20,pady=10)
tree.heading('Description',text='Description')
tree.heading('Qty',text='Qty')
tree.heading('Price',text='Price')
tree.heading('Total',text='Total')


save_btn = tkinter.Button(frame,text="          Save          ",command=save)
save_btn.grid(row=5,column=0,columnspan=3)


print_btn = tkinter.Button(frame,text="          Print         ")
print_btn.grid(row=5,column=3,columnspan=3)

newI = tkinter.Button(frame,text="Generate New Invoice",command=newI)
newI.grid(row=6,column=0,columnspan=6,sticky='news',pady=10,padx=15)



window.mainloop()
